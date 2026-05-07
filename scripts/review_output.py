#!/usr/bin/env python3
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl, re, sqlite3, random
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent / "core"))
import config
from ingest import load_glossary

wb = openpyxl.load_workbook(config.OUTPUT_DIR / "0128_liyuan_zh-CN_optimized.xlsx")
ws = wb.active
headers = [c.value for c in ws[1]]
rows = list(ws.iter_rows(min_row=2, values_only=True))
print("=" * 50)
print("Agent 输出审查报告")
print("=" * 50)
print(f"文件: 0128_liyuan_zh-CN_optimized.xlsx")
print(f"总行数: {len(rows)}")

# 1. 列完整性
expected = ['row_num', 'source', 'draft', 'optimized', 'change_type', 'change_reason', 'locked', 'notes']
missing = [h for h in expected if h not in headers]
print(f"\n[1] 列完整性: {'PASS' if not missing else 'FAIL - missing: ' + str(missing)}")

# 2. 空值扫描
opt_idx = headers.index('optimized')
locked_idx = headers.index('locked')
empty_optimized = [i+2 for i, r in enumerate(rows) if not r[opt_idx] and not r[locked_idx]]
print(f"[2] 空值扫描: {len(empty_optimized)} 行 optimized 为空（非 locked）")
if empty_optimized:
    print(f"    行号: {empty_optimized[:10]}")

# 3. 标签完整性
tag_pattern = re.compile(r'<(color|i|b|size|font)[^>]*>')
close_pattern = re.compile(r'</(color|i|b|size|font)>')
src_idx = headers.index('source')
opt_idx = headers.index('optimized')
tag_issues = []
for i, r in enumerate(rows):
    for text_idx, col_name in [(src_idx, 'source'), (opt_idx, 'optimized')]:
        text = str(r[text_idx] or '')
        tags = tag_pattern.findall(text)
        for tag in set(tags):
            opens = text.count(f'<{tag}')
            closes = len(close_pattern.findall(text))
            if opens != closes:
                tag_issues.append((i+2, col_name, tag))
                break
print(f"[3] 标签完整性: {len(tag_issues)} 处标签不匹配")
for t in tag_issues[:5]:
    print(f"    Row {t[0]} {t[1]}: <{t[2]}> 不成对")

# 4. change_type 合理性 (结合 draft==translation)
draft_idx = headers.index('draft')
opt_idx = headers.index('optimized')
ctype_idx = headers.index('change_type')
valid_types = {'terminology_fix', 'style_alignment', 'grammar_fix', 'polish', 'no_change', ''}
ctypes = [r[ctype_idx] for r in rows if r[ctype_idx]]
no_change_cnt = ctypes.count('no_change')
no_change_ratio = no_change_cnt / len(ctypes) if ctypes else 0
invalid_types = [c for c in set(ctypes) if c not in valid_types]

# Count real no_change (draft == optimized)
real_no = sum(1 for r in rows if r[ctype_idx] == 'no_change' and str(r[draft_idx] or '') == str(r[opt_idx] or ''))
false_neg = sum(1 for r in rows if r[ctype_idx] == 'no_change' and str(r[draft_idx] or '') != str(r[opt_idx] or ''))
false_pos = sum(1 for r in rows if r[ctype_idx] != 'no_change' and str(r[draft_idx] or '') == str(r[opt_idx] or ''))

print(f"[4] change_type 合理性:")
print(f"    no_change: {no_change_cnt}/{len(ctypes)} ({no_change_ratio:.1%})")
print(f"      real (draft==opt): {real_no}, false negative: {false_neg}")
print(f"    false positive (draft==opt but ctype!=no_change): {false_pos}")
print(f"    invalid types: {invalid_types if invalid_types else 'None'}")
if no_change_ratio > 0.30 and false_neg > 5:
    print("    WARN: no_change 比例 > 30% 且存在大量 false negative")

# 5. 术语一致性 (随机 5%, 支持大小写/冠词变体)
glossary = load_glossary(str(config.BASE_DIR / 'input' / 'glossary_deduped.xlsx'))
sample = random.sample(rows, min(int(len(rows)*0.05)+1, 30))
term_miss = 0
miss_details = []
for r in sample:
    src = str(r[src_idx] or '')
    opt = str(r[opt_idx] or '')
    opt_lower = opt.lower()
    for st, tt in glossary:
        if st not in src:
            continue
        # Variants: with/without leading "the ", case-insensitive
        variants = {tt.lower()}
        if tt.lower().startswith("the "):
            variants.add(tt.lower()[4:])
        else:
            variants.add("the " + tt.lower())
        if not any(v in opt_lower for v in variants):
            term_miss += 1
            miss_details.append((r[0], st, tt))
            break
print(f"[5] 术语一致性: 抽检 {len(sample)} 行, {term_miss} 处偏差")
for d in miss_details[:5]:
    print(f"    Row {d[0]}: 术语 '{d[1]}' -> 期望 '{d[2]}'")

# 6. 风格一致性 (首尾对比)
first_opt = str(rows[0][opt_idx] or '')
last_opt = str(rows[-1][opt_idx] or '')
print(f"[6] 风格一致性: 首尾段 tone 抽查")
print(f"    首段: {first_opt[:80]}")
print(f"    末段: {last_opt[:80]}")

print("\n" + "=" * 50)
print("审查结论:", "通过" if (not missing and not empty_optimized and not invalid_types and term_miss == 0) else "需修正")
print("=" * 50)

wb.close()
