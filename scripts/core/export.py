"""
export.py — Export final results to xlsx/csv.
Mode A: source | translation | locked
Mode B: source | draft | optimized(富文本差异) | locked
         + change_type / change_reason (AI-generated, batched)
"""

import csv
import re
import json
import difflib
from pathlib import Path
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles.colors import Color
import sqlite3

import config
WORKSPACE_DB = str(config.WORKSPACE_DIR / "workspace.db")

# ---------------------------------------------------------------------------
# Rich-text diff
# ---------------------------------------------------------------------------

def _make_rich_diff(draft: str, optimized: str) -> CellRichText | str:
    """Return optimized text with differences from draft highlighted in red bold.
    Reference: AIPE _diff_bold logic.
    """
    if not draft or not optimized or draft.strip() == optimized.strip():
        return optimized

    draft_tokens = re.split(r'(\s+)', draft)
    opt_tokens = re.split(r'(\s+)', optimized)
    sm = difflib.SequenceMatcher(None, draft_tokens, opt_tokens, autojunk=False)
    raw_blocks, has_change = [], False

    red_bold = InlineFont(b=True)
    red_bold.color = Color(rgb="FF0000")

    for tag, _i1, _i2, j1, j2 in sm.get_opcodes():
        chunk = ''.join(opt_tokens[j1:j2])
        if not chunk:
            continue
        if tag == 'equal':
            raw_blocks.append(chunk)
        else:
            has_change = True
            raw_blocks.append(TextBlock(red_bold, chunk))

    if not has_change:
        return optimized

    # Merge standalone whitespace blocks into neighbours to avoid Excel dropping them
    blocks = []
    for block in raw_blocks:
        is_ws = isinstance(block, str) and block.strip() == ''
        if is_ws:
            if blocks:
                prev = blocks[-1]
                if isinstance(prev, str):
                    blocks[-1] = prev + block
                else:
                    blocks[-1] = TextBlock(prev.font, prev.text + block)
            else:
                blocks.append(block)
        else:
            if blocks and isinstance(blocks[-1], str) and blocks[-1].strip() == '':
                ws = blocks.pop()
                if isinstance(block, str):
                    blocks.append(ws + block)
                else:
                    blocks.append(TextBlock(block.font, ws + block.text))
            else:
                blocks.append(block)

    return CellRichText(blocks) if blocks else optimized


# ---------------------------------------------------------------------------
# Change classification (read from DB, written during translate)
# ---------------------------------------------------------------------------

def classify_rows(mode: str = "optimize") -> dict:
    """Read change classifications from DB (populated during translate)."""
    conn = sqlite3.connect(WORKSPACE_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT row_num, change_type, change_reason FROM rows WHERE change_type != '' OR change_reason != '' ORDER BY row_num"
    ).fetchall()
    conn.close()
    return {str(r["row_num"]): {"change_type": r["change_type"], "change_reason": r["change_reason"]} for r in rows}


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_xlsx(output_path: str, mode: str = "new"):
    """Create standardized output xlsx from workspace DB."""
    conn = sqlite3.connect(WORKSPACE_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT row_num, source, draft, translation, locked, notes FROM rows ORDER BY row_num"
    ).fetchall()
    conn.close()

    # Pre-compute classifications for Mode B
    classifications = {}
    if mode == "optimize":
        classifications = classify_rows(mode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Final"

    if mode == "new":
        headers = ["row_num", "source", "translation", "locked"]
        ws.append(headers)
        for r in rows:
            ws.append([
                r["row_num"], r["source"], r["translation"] or "",
                r["locked"],
            ])
    else:
        headers = ["row_num", "source", "draft", "optimized", "change_type", "change_reason", "locked", "notes"]
        ws.append(headers)
        for r in rows:
            draft = r["draft"] or ""
            optimized = r["translation"] or ""
            rich = _make_rich_diff(draft, optimized)
            cls = classifications.get(str(r["row_num"]), {})
            ws.append([
                r["row_num"], r["source"], draft, "",
                cls.get("change_type", "no_change"),
                cls.get("change_reason", ""),
                r["locked"],
                r["notes"] or "",
            ])
            cell = ws.cell(row=ws.max_row, column=4)
            cell.value = rich

    wb.save(output_path)
    print(f"[export] Saved -> {output_path}")


def export_csv(output_path: str):
    conn = sqlite3.connect(WORKSPACE_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT row_num, source, draft, translation, status FROM rows ORDER BY row_num"
    ).fetchall()
    conn.close()

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["row_num", "source", "draft", "translation", "status"])
        for r in rows:
            writer.writerow([r["row_num"], r["source"], r["draft"], r["translation"], r["status"]])
    print(f"[export] Saved -> {output_path}")
