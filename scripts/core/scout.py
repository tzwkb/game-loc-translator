"""
scout.py — Context Scout: analyze input file and generate structured context report.
"""

import json
import sqlite3
from pathlib import Path

import config
from engine import _call_api_raw, load_bilingual_file, get_file_headers


def _build_scout_prompt(input_file: str, glossary: list, kb: list, sample_rows: list) -> list:
    """Build messages for LLM context analysis."""
    headers = get_file_headers(input_file)
    header_str = " | ".join(f"{idx}:{name}" for idx, name in headers)

    glossary_preview = "\n".join(f"  - {st} → {tt}" for st, tt in glossary[:30]) if glossary else "  (empty)"
    kb_preview = "\n".join(f"  [{e.get('category','')}]: {e.get('text','')[:100]}" for e in kb[:10]) if kb else "  (empty)"

    sample_text = "\n".join(f"  {i+1}. {s[:120]}" for i, s in enumerate(sample_rows[:20]))

    system = (
        "你是游戏本地化项目的 Context Scout（上下文侦察员）。"
        "你的任务：分析输入文件的题材、IP、语调、风格特征，识别风险点和关键术语，"
        "输出结构化的上下文报告，供 Translator 使用。"
    )

    user = (
        f"输入文件: {Path(input_file).name}\n"
        f"表头: {header_str}\n\n"
        f"术语表预览（前30条）:\n{glossary_preview}\n\n"
        f"知识库预览（前10条）:\n{kb_preview}\n\n"
        f"源语言样本（前20行）:\n{sample_text}\n\n"
        "请输出 JSON 格式的上下文报告，字段如下:\n"
        '{\n'
        '  "genre": "题材类型，如武侠RPG/科幻SLG/奇幻ACT",\n'
        '  "tone": "整体语调，如文艺与口语并存/正式/诙谐",\n'
        '  "style_anchor": "风格基调描述（50字内），供Translator作为prompt使用",\n'
        '  "key_terms": {"术语": "翻译要求或注意点"},\n'
        '  "risk_flags": ["风险描述，如文化梗/敏感词/谐音"],\n'
        '  "placeholder_rules": ["占位符处理规则，如保留{}变量"],\n'
        '  "length_constraints": "是否有max_length限制及处理建议",\n'
        '  "gender_notes": "gender列的处理要求",\n'
        '  "notes": "其他观察"\n'
        '}\n'
        "仅输出合法 JSON，不要 markdown 代码块，不要解释。"
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def run_scout(input_file: str, glossary_file: str = None, kb_file: str = None) -> dict:
    """Run context scout and write report to workspace/scout_report.json."""
    from ingest import load_glossary, load_knowledge_base

    # Load glossary
    glossary = []
    if glossary_file:
        glossary = load_glossary(glossary_file)

    # Load knowledge base
    kb = []
    if kb_file:
        kb = load_knowledge_base(kb_file)

    # Sample rows from input
    sample_rows = []
    path = Path(input_file)
    if path.suffix.lower() == ".csv":
        import csv
        with open(input_file, encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if row:
                    sample_rows.append(row[0] if row[0] else "")
    else:
        from openpyxl import load_workbook
        wb = load_workbook(input_file, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                sample_rows.append(str(row[0]))
        wb.close()

    # Call API
    msgs = _build_scout_prompt(input_file, glossary, kb, sample_rows)
    raw = _call_api_raw(msgs)
    if not raw:
        print("[scout] API call failed.")
        return {}

    try:
        report = json.loads(raw.strip().strip("```json").strip("```").strip())
    except json.JSONDecodeError:
        print(f"[scout] JSON parse failed. Raw:\n{raw[:500]}")
        return {}

    # Write report
    report_path = config.WORKSPACE_DIR / "scout_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[scout] Report -> {report_path}")
    return report
