"""
engine.py — Core processing engine.
Handles: API client, async batch calls, prompt loading, TM embedding index, glossary matching.
"""

import asyncio
import datetime
import json
import os
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

try:
    from openai import OpenAI, AsyncOpenAI
except ImportError:
    raise ImportError("openai package not found. Run: pip install openai openpyxl sentence-transformers faiss-cpu")

import config

# ---------------------------------------------------------------------------
# Prompt loading
# ---------------------------------------------------------------------------
_prompt_cache: dict = {}

def load_prompt(name: str, project_id: str = "") -> str:
    """Load prompt by name. For 'translate'/'optimize', auto-append project file if exists."""
    cache_key = f"{name}:{project_id}"
    if cache_key not in _prompt_cache:
        base_path = config.PROMPTS_DIR / f"{name}_base.md"
        base = base_path.read_text(encoding="utf-8") if base_path.exists() else ""
        if project_id:
            proj_path = config.PROMPTS_DIR / f"{name}_project_{project_id}.md"
            if proj_path.exists():
                base += "\n\n" + proj_path.read_text(encoding="utf-8")
        _prompt_cache[cache_key] = base.strip()
    return _prompt_cache[cache_key]

# ---------------------------------------------------------------------------
# API logging
# ---------------------------------------------------------------------------
_log_lock = threading.Lock()
_log_file = None

def init_api_log():
    global _log_file
    config.LOG_DIR.mkdir(exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = config.LOG_DIR / f"api_{ts}.jsonl"
    _log_file = open(log_path, "a", encoding="utf-8", buffering=1)
    return str(log_path)

def _log_api_call(messages: list, response: str | None, error: str | None = None, attempt: int = 0):
    if _log_file is None:
        return
    record = {
        "ts": datetime.datetime.now().isoformat(),
        "model": config.MODEL,
        "attempt": attempt,
        "messages": messages,
        "response": response,
    }
    if error:
        record["error"] = error
    with _log_lock:
        _log_file.write(json.dumps(record, ensure_ascii=False) + "\n")

# ---------------------------------------------------------------------------
# API clients
# ---------------------------------------------------------------------------
_client: OpenAI | None = None
_async_client: AsyncOpenAI | None = None

def _get_client() -> OpenAI:
    global _client
    if _client is None:
        _client = OpenAI(api_key=config.API_KEY, base_url=config.API_BASE_URL)
    return _client

def _get_async_client() -> AsyncOpenAI:
    global _async_client
    if _async_client is None:
        _async_client = AsyncOpenAI(api_key=config.API_KEY, base_url=config.API_BASE_URL)
    return _async_client

def strip_json_fences(text: str) -> str:
    text = re.sub(r'^```[a-z]*\n?', '', text.strip())
    return re.sub(r'\n?```$', '', text).strip()

# ---------------------------------------------------------------------------
# Sync API call with retry
# ---------------------------------------------------------------------------
def _call_api_raw(messages: list, temperature: float = None, attempt: int = 0) -> str | None:
    temp = temperature if temperature is not None else config.TEMPERATURE
    try:
        resp = _get_client().chat.completions.create(
            model=config.MODEL, messages=messages, max_tokens=config.MAX_TOKENS,
            temperature=temp, timeout=config.REQUEST_TIMEOUT,
        )
        text = resp.choices[0].message.content.strip()
        _log_api_call(messages, text, attempt=attempt)
        return text
    except Exception as exc:
        if attempt < config.MAX_RETRIES:
            wait = config.RETRY_DELAY * (2 ** attempt)
            time.sleep(wait)
            return _call_api_raw(messages, temp, attempt + 1)
        _log_api_call(messages, None, error=str(exc), attempt=attempt)
        return None

# ---------------------------------------------------------------------------
# Async API call with retry
# ---------------------------------------------------------------------------
class _null_sem:
    async def __aenter__(self): return self
    async def __aexit__(self, *_): pass

_NULL_SEM = _null_sem()

async def _async_call_api_raw(messages: list, temperature: float = None,
                               semaphore: asyncio.Semaphore = None, attempt: int = 0) -> str | None:
    temp = temperature if temperature is not None else config.TEMPERATURE
    async with (semaphore or _NULL_SEM):
        try:
            resp = await _get_async_client().chat.completions.create(
                model=config.MODEL, messages=messages, max_tokens=config.MAX_TOKENS,
                temperature=temp, timeout=config.REQUEST_TIMEOUT,
            )
            text = resp.choices[0].message.content.strip()
            _log_api_call(messages, text, attempt=attempt)
            return text
        except Exception as exc:
            if attempt < config.MAX_RETRIES:
                wait = config.RETRY_DELAY * (2 ** attempt)
                await asyncio.sleep(wait)
                return await _async_call_api_raw(messages, temp, semaphore, attempt + 1)
            _log_api_call(messages, None, error=str(exc), attempt=attempt)
            return None

# ---------------------------------------------------------------------------
# Batch message builder
# ---------------------------------------------------------------------------
def build_batch_messages(batch_items: list, mode: str, project_id: str = "",
                         style_anchor: str = "", glossary_hints: list = None,
                         kb_snippets: list = None, rag_refs: list = None) -> list:
    """Build messages for a batch API call.
    batch_items: list of dicts with keys: label, source, draft(optional), key
    mode: 'translate' or 'optimize'
    """
    system_content = load_prompt(mode, project_id)
    if style_anchor:
        system_content += f"\n\n[风格锚定] 全文基调：{style_anchor}"

    preamble_parts = []
    if glossary_hints:
        lines = "\n".join(f"  - {s} → {t}" for s, t in glossary_hints)
        preamble_parts.append(f"强制术语表（必须严格遵循）：\n{lines}")
    if kb_snippets:
        lines = "\n".join(f"  [{note}]: {text}" for note, text in kb_snippets)
        preamble_parts.append(f"项目知识库参考：\n{lines}")
    if rag_refs:
        lines = "\n".join(f"  [历史译法]: {s} → {t}" for s, t in rag_refs)
        preamble_parts.append(f"RAG 语料参考（仅供参考，不强制）：\n{lines}")

    row_blocks = []
    for item in batch_items:
        label = item["label"]
        src = item["source"]
        key = item.get("key", "")
        parts = []
        if key:
            parts.append(f"Key: {key}")
        parts.append(f"SOURCE: {src}")
        if mode == "optimize" and item.get("draft"):
            parts.append(f"DRAFT: {item['draft']}")
        row_blocks.append(f"[{label}]\n" + "\n".join(parts))

    if mode == "optimize":
        instructions = (
            f"请处理以下 {len(batch_items)} 行文本。"
            + "对每行初译进行优化，输出优化后的译文。"
            + "\n返回 JSON 对象，键为行号（字符串），值为包含以下字段的对象："
            + '\n  {"translation": "译文", "change_type": "类型", "change_reason": "原因"}'
            + "\nchange_type 必须从以下选项中选择：terminology_fix, style_alignment, grammar_fix, polish, no_change"
            + "\nchange_reason 用中文一句话简要说明修改原因，尽量简洁（15字以内）。"
            + '\n示例: {"1": {"translation": "译文1", "change_type": "style_alignment", "change_reason": "调整语气更自然"}, "2": {...}}'
            + "\n仅输出合法 JSON，不要 markdown 代码块，不要解释。"
        )
    else:
        instructions = (
            f"请处理以下 {len(batch_items)} 行文本。"
            + "输出每行的翻译译文。"
            + "\n返回 JSON 对象，键为行号（字符串），值为译文。"
            + '\n示例: {"1": "译文1", "2": "译文2"}'
            + "\n仅输出合法 JSON，不要 markdown 代码块，不要解释。"
        )

    user_parts = list(preamble_parts)
    user_parts.append(instructions)
    user_parts.append("--- ROWS ---")
    user_parts.extend(row_blocks)
    user_parts.append("--- END ---")

    return [
        {"role": "system", "content": system_content},
        {"role": "user", "content": "\n\n".join(user_parts)},
    ]


def _parse_batch_response(raw: str, labels: list) -> dict:
    """Parse API response. Supports both legacy string values and new dict values.
    Returns {label: {"translation": str, "change_type": str, "change_reason": str}}.
    """
    if not raw:
        return {lbl: {"translation": None, "change_type": None, "change_reason": None} for lbl in labels}
    label_set = set(labels)
    result = {}
    # Try JSON first
    try:
        data = json.loads(strip_json_fences(raw))
        if isinstance(data, dict):
            for k, v in data.items():
                try:
                    lbl = int(k)
                    if lbl in label_set:
                        if isinstance(v, dict):
                            result[lbl] = {
                                "translation": v.get("translation") if isinstance(v.get("translation"), str) else None,
                                "change_type": v.get("change_type") if isinstance(v.get("change_type"), str) else None,
                                "change_reason": v.get("change_reason") if isinstance(v.get("change_reason"), str) else None,
                            }
                        elif isinstance(v, str):
                            # Legacy format
                            result[lbl] = {"translation": v, "change_type": None, "change_reason": None}
                except (ValueError, TypeError):
                    pass
    except (json.JSONDecodeError, ValueError, AttributeError):
        pass

    # Fallback: [N] pattern
    if not result:
        parts = re.split(r'(?m)^\[(\d+)\]\s*\n?', raw)
        i = 1
        while i + 1 < len(parts):
            try:
                lbl = int(parts[i])
                content = parts[i + 1].strip()
                if lbl in label_set:
                    result[lbl] = {"translation": content, "change_type": None, "change_reason": None}
            except (ValueError, IndexError):
                pass
            i += 2

    for lbl in labels:
        if lbl not in result:
            result[lbl] = {"translation": None, "change_type": None, "change_reason": None}
    return result


async def translate_batch(batch_items: list, mode: str = "translate", project_id: str = "",
                          style_anchor: str = "",
                          glossary_hints: list = None, kb_snippets: list = None,
                          rag_refs: list = None, semaphore: asyncio.Semaphore = None) -> dict:
    """Async batch translate/optimize. Returns {label: text_or_None}.
    If response is truncated due to max_tokens, key '_truncated' is set True.
    """
    labels = [item["label"] for item in batch_items]
    msgs = build_batch_messages(batch_items, mode, project_id, style_anchor,
                                glossary_hints, kb_snippets, rag_refs)
    raw = await _async_call_api_raw(msgs, semaphore=semaphore)
    if raw is None:
        return {lbl: None for lbl in labels}
    # Detect systematic truncation (max_tokens insufficient)
    stripped = strip_json_fences(raw).strip()
    truncated = bool(stripped and not (stripped.endswith('}') or stripped.endswith(']')))
    result = _parse_batch_response(raw, labels)
    if truncated:
        result["_truncated"] = True
    return result

# ---------------------------------------------------------------------------
# TM / Embedding Index
# ---------------------------------------------------------------------------
class EmbeddingIndex:
    MODEL_NAME = "paraphrase-multilingual-MiniLM-L12-v2"

    def __init__(self, entries: list, cache_path: Path = None):
        os.environ.setdefault("HF_HUB_OFFLINE", "1")
        os.environ.setdefault("TRANSFORMERS_OFFLINE", "1")
        self.entries = entries  # list of (source, target)
        self._query_cache: dict = {}
        import numpy as np
        self._np = np

        cached = None
        if cache_path and cache_path.exists():
            cached = np.load(str(cache_path))
        cache_valid = cached is not None and cached.shape[0] == len(entries)

        from sentence_transformers import SentenceTransformer
        if cache_valid:
            self._matrix = cached
            self._model = SentenceTransformer(self.MODEL_NAME)
        else:
            self._model = SentenceTransformer(self.MODEL_NAME)
            self._matrix = self._model.encode(
                [s for s, _ in entries], batch_size=256,
                normalize_embeddings=True, show_progress_bar=False, convert_to_numpy=True,
            )
            if cache_path:
                np.save(str(cache_path), self._matrix)

    def query(self, text: str, top_k: int = 3, threshold: float = 0.55) -> list:
        if not text or not text.strip():
            return []
        effective_threshold = 0.90 if len(text.strip()) < 10 else threshold
        vec = self._query_cache.get(text)
        if vec is None:
            vec = self._model.encode([text], normalize_embeddings=True, convert_to_numpy=True)[0]
            self._query_cache[text] = vec
        sims = self._matrix.dot(vec)
        k = min(top_k, len(sims))
        top_idx = self._np.argpartition(sims, -k)[-k:].tolist()
        results = [(float(sims[i]), self.entries[i][0], self.entries[i][1])
                   for i in top_idx if float(sims[i]) >= effective_threshold]
        results.sort(reverse=True)
        return results


def build_tm_index(entries: list, source_path: str = None) -> tuple:
    cache_path = None
    if source_path:
        p = Path(source_path)
        cache_path = p.parent / (p.stem + ".embeddings.npy")
    idx = EmbeddingIndex(entries, cache_path=cache_path)
    return idx, f"语义向量索引（dim={idx._matrix.shape[1]}）"

# ---------------------------------------------------------------------------
# Glossary matching
# ---------------------------------------------------------------------------
def find_matching_terms(text: str, glossary: list) -> list:
    if not text or not glossary:
        return []
    return [(st, tt) for st, tt in glossary if st in text]

# ---------------------------------------------------------------------------
# Excel utilities
# ---------------------------------------------------------------------------
def get_cell_text(cell) -> str:
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    try:
        parts = [item.text if hasattr(item, "text") else str(item) for item in val]
        return "".join(parts)
    except TypeError:
        return str(val)


def load_bilingual_file(filepath: str, st_col: int, tt_col: int) -> list:
    path = Path(filepath)
    rows = []
    if path.suffix.lower() == ".csv":
        import csv
        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            for row in csv.reader(f):
                if len(row) >= max(st_col, tt_col):
                    st = row[st_col - 1].strip()
                    tt = row[tt_col - 1].strip()
                    if st and tt:
                        rows.append((st, tt))
    else:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        for row in wb.active.iter_rows(values_only=True):
            if row and len(row) >= max(st_col, tt_col):
                st = str(row[st_col - 1] or "").strip()
                tt = str(row[tt_col - 1] or "").strip()
                if st and tt:
                    rows.append((st, tt))
        wb.close()
    return rows


def get_file_headers(filepath: str) -> list:
    path = Path(filepath)
    if path.suffix.lower() == ".csv":
        import csv
        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            row = next(csv.reader(f), [])
        return [(i + 1, v.strip() or f"列{i+1}") for i, v in enumerate(row) if v.strip()]

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(max_row=3, values_only=True):
        non_empty = [(i, v) for i, v in enumerate(row) if v is not None]
        if non_empty:
            last = non_empty[-1][0]
            headers = [(i + 1, str(row[i]) if row[i] is not None else f"列{i+1}")
                       for i in range(last + 1)]
            break
    wb.close()
    return headers
